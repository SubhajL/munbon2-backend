"""
Wave 2.1b — strict SCADA workbook generator (scripts/build_scada_config.py).

The generator replaces scripts/excel_to_canal_sections.py (whose Sheet1
`ระยะทาง (เมตร)` column actually holds KILOMETRES — the 1000x length bug) with a
fail-closed pipeline: Sheet1 gate positions (per-gate `km`) cut the
Characteristics survey rows (canal + chainage spans, true metres) into
gate-to-gate edge segments in the 2.1a multi-segment geometry model, plus a
per-edge coverage report. Every artifact is versioned with the workbook SHA-256,
and the committed artifacts are locked to regeneration (no hand-edit drift).

Semantics locked here (documented in the generator):
- Gate position authority is Sheet1's `km` column; ระยะทาง is a validated
  cross-check against the next existing physical valve.
- Segment q_max is the Characteristics design discharge `Qd` (complete on every
  hydraulic row); `Qmax (จากแผนรอบเวรส่งน้ำ)` is a rotation-plan operational
  value and is patchy (absent for the whole RMC group).
- side_slope is DERIVED from the design flow area: z = (A/D - B)/D. The old
  file's 0.06 was the `t` column — the 6 cm concrete lining THICKNESS misread
  as a slope. The derivation yields 1.5 on all 99 survey rows.
- Survey rows that cross a gate position are split at the gate; chainage
  arithmetic only, lengths conserved (every row's length equals its chainage
  delta — validated, fail-closed).
- Rows that cannot be emitted (the Flume row has no cross-section; the 7R tail
  lies beyond the last gate) are REPORTED, never defaulted. Sill/control
  structures are RID-gated and never emitted.
"""
import hashlib
import importlib.util
import json
from pathlib import Path

import pytest

from core.config_loader import (
    load_canal_geometry_config,
    load_network_config,
)
from core.conveyance_loss import parse_chainage_m, sections_by_edge_from_geometry
from core.network_topology import edges_from_names

SERVICE_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = SERVICE_ROOT.parents[1]
WORKBOOK = (
    SERVICE_ROOT
    / "data"
    / "sources"
    / "scada"
    / "SCADA Section Detailed Information 2026-07-14 V3.0 SL.xlsx"
)
WORKBOOK_SHA256 = "528a3fe3978e916ce2048189239045c9ecae5d74f456a2100c9c946ca2787e1c"
CONFIG_DIR = SERVICE_ROOT / "src" / "config"

_SPEC = importlib.util.spec_from_file_location(
    "build_scada_config", SERVICE_ROOT / "scripts" / "build_scada_config.py"
)
bsc = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(bsc)


class TestParseKmMarker:
    @pytest.mark.parametrize(
        "text,metres",
        [
            ("0+000", 0.0),
            ("0+300", 300.0),
            ("6+880", 6880.0),
            ("39+050", 39050.0),
            ("1+620.5", 1620.5),
        ],
    )
    def test_parses_km_plus_metres(self, text, metres):
        assert bsc.parse_km_marker(text) == metres

    @pytest.mark.parametrize("bad", [None, "", "abc", "6-880", "6+", "+880", 12.5])
    def test_rejects_unparseable_markers(self, bad):
        # The extractor must fail closed: a silent None here becomes a silently
        # unmapped survey row downstream.
        with pytest.raises(bsc.WorkbookError):
            bsc.parse_km_marker(bad)

    @pytest.mark.parametrize(
        "metres,text",
        [
            (0.0, "0+000"),
            (300.0, "0+300"),
            (50.0, "0+050"),
            (6880.0, "6+880"),
            (39050.0, "39+050"),
        ],
    )
    def test_formats_zero_padded_markers(self, metres, text):
        assert bsc.format_km_marker(metres) == text

    @pytest.mark.parametrize("metres", [0.0, 130.0, 6880.0, 39050.0, 1620.5])
    def test_round_trips_through_the_runtime_parser(self, metres):
        # The emitted from_km/to_km strings are consumed by
        # core.conveyance_loss.parse_chainage_m — the two parsers must agree.
        assert parse_chainage_m(bsc.format_km_marker(metres)) == metres


class TestDeriveSideSlope:
    def test_recovers_the_trapezoid_side_slope_from_design_area(self):
        # LMC 0+170-1+620: A=11.609, B=3.5, D=1.85 -> z = (A/D - B)/D = 1.5
        assert bsc.derive_side_slope(11.609, 3.5, 1.85) == 1.5

    def test_exact_trapezoid_round_trip(self):
        b, d, z = 2.4, 1.55, 1.5
        area = d * (b + z * d)
        assert bsc.derive_side_slope(area, b, d) == z

    def test_rectangular_section_yields_zero(self):
        assert bsc.derive_side_slope(2.0 * 1.5, 2.0, 1.5) == 0.0

    def test_rejects_negative_slope(self):
        # An area smaller than the rectangle B*D means the inputs are
        # inconsistent — never emit a negative slope.
        with pytest.raises(bsc.WorkbookError):
            bsc.derive_side_slope(1.0, 2.0, 1.5)

    def test_rejects_implausibly_flat_slope(self):
        with pytest.raises(bsc.WorkbookError):
            bsc.derive_side_slope(100.0, 1.0, 1.0)  # z = 99

    @pytest.mark.parametrize("area,b,d", [(1.0, 0.0, 1.0), (1.0, 1.0, 0.0)])
    def test_rejects_nonpositive_dimensions(self, area, b, d):
        with pytest.raises(bsc.WorkbookError):
            bsc.derive_side_slope(area, b, d)


class TestLiningFromNote:
    @pytest.mark.parametrize(
        "note,lining",
        [
            ("คลองดาดคอนกรีต", "concrete"),
            ("คลองคาดคอนกรีต", "concrete"),  # the 5R-4L-38R-LMC typo spelling
            ("คลองดิน", "earth"),
            (None, None),
            ("Hl=3.00-6.00", None),  # a remark, not a lining statement
            ("Flume", None),
        ],
    )
    def test_maps_survey_notes_to_lining(self, note, lining):
        assert bsc.lining_from_note(note) == lining


GATES_SYNTH = [
    "M(0,0)",
    "M(0,1)",
    "M(0,2)",
    "M(0,1;1,0)",
    "M(0,1;1,1)",
    "M(0,1;1,1;1,0)",
    "M(0,1;1,1;2,0)",
]


class TestBuildSerialChains:
    def test_groups_gates_into_serial_chains(self):
        chains = bsc.build_serial_chains(GATES_SYNTH)
        assert sorted(chains, key=len, reverse=True)[0] == [
            "M(0,0)",
            "M(0,1)",
            "M(0,2)",
        ]
        assert ["M(0,1;1,0)", "M(0,1;1,1)"] in chains

    def test_distinct_branch_indexes_are_distinct_chains(self):
        # 7R (branch 1) and 7L (branch 2) hang off the same parent gate but are
        # different canals — they must never merge into one chainage axis.
        chains = bsc.build_serial_chains(GATES_SYNTH)
        assert ["M(0,1;1,1;1,0)"] in chains
        assert ["M(0,1;1,1;2,0)"] in chains

    def test_every_gate_lands_in_exactly_one_chain(self):
        chains = bsc.build_serial_chains(GATES_SYNTH)
        flat = [g for chain in chains for g in chain]
        assert sorted(flat) == sorted(GATES_SYNTH)

    def test_chain_order_follows_serial_position_not_input_order(self):
        chains = bsc.build_serial_chains(["M(0,2)", "M(0,0)", "M(0,1)"])
        assert chains == [["M(0,0)", "M(0,1)", "M(0,2)"]]

    def test_chain_preserves_sparse_valve_numbers_without_inventing_a_gate(self):
        chains = bsc.build_serial_chains(["M(0,2)", "M(0,0)"])
        assert chains == [["M(0,0)", "M(0,2)"]]

    def test_sparse_chain_without_position_zero_fails_closed(self):
        with pytest.raises(bsc.WorkbookError, match="must start at position 0"):
            bsc.build_serial_chains(["M(0,2)"])

    def test_duplicate_serial_position_fails_closed(self):
        with pytest.raises(bsc.WorkbookError, match="duplicate serial position"):
            bsc.build_serial_chains(["M(0,0)", "M (0,0)"])


class TestValidateGateDistances:
    CHAIN = [["M(0,0)", "M(0,2)"]]

    def test_accepts_distance_to_the_next_existing_sparse_valve(self):
        gates = {
            "M(0,0)": {"km_m": 0.0, "distance_km": 1.62},
            "M(0,2)": {"km_m": 1620.0, "distance_km": None},
        }
        assert bsc.validate_gate_distances(self.CHAIN, gates) is None

    def test_rejects_distance_left_at_the_removed_valve_position(self):
        gates = {
            "M(0,0)": {"km_m": 0.0, "distance_km": 0.17},
            "M(0,2)": {"km_m": 1620.0, "distance_km": None},
        }
        with pytest.raises(
            bsc.WorkbookError,
            match=r"M\(0,0\).*0\.17 km.*1\.62 km.*M\(0,2\)",
        ):
            bsc.validate_gate_distances(self.CHAIN, gates)


class TestMatchChainsToSurvey:
    CHAINS = [["M(0,0)", "M(0,1)", "M(0,2)"], ["M(0,1;1,0)", "M(0,1;1,1)"]]
    CANALS = {
        "M(0,0)": "Outlet",
        "M(0,1)": None,
        "M(0,2)": "LMC",
        "M(0,1;1,0)": "RMC",
        "M(0,1;1,1)": "RMC",
    }

    def test_matches_survey_canals_by_member_gate_names(self):
        matched = bsc.match_chains_to_survey(self.CHAINS, self.CANALS, ["LMC", "RMC"])
        assert matched["LMC"] == self.CHAINS[0]
        assert matched["RMC"] == self.CHAINS[1]

    def test_survey_canal_with_no_chain_fails_closed(self):
        # A surveyed canal that maps to no gate chain would silently drop its
        # whole geometry — that is exactly the drift this generator forbids.
        with pytest.raises(bsc.WorkbookError):
            bsc.match_chains_to_survey(self.CHAINS, self.CANALS, ["LMC", "9R-LMC"])

    def test_ambiguous_name_across_two_chains_fails_closed(self):
        canals = dict(self.CANALS, **{"M(0,1;1,1;1,0)": "RMC"})
        chains = self.CHAINS + [["M(0,1;1,1;1,0)"]]
        with pytest.raises(bsc.WorkbookError):
            bsc.match_chains_to_survey(chains, canals, ["RMC"])


def _row(from_m, to_m, **over):
    row = {
        "canal": "LMC",
        "from_m": float(from_m),
        "to_m": float(to_m),
        "length_m": float(to_m - from_m),
        "qd": 5.0,
        "area_m2": 7.84,
        "manning_n": 0.018,
        "bed_slope": 0.0002,
        "bottom_width_m": 2.5,
        "depth_m": 1.6,
        "note": "คลองดาดคอนกรีต",
        "excel_row": 99,
    }
    row.update(over)
    return row


class TestAssignRowsToEdges:
    POSITIONS = [("M(0,0)", 0.0), ("M(0,1)", 300.0), ("M(0,2)", 1620.0)]

    def test_exact_boundary_rows_map_one_to_one(self):
        pieces, skipped, coverage = bsc.assign_rows_to_edges(
            self.POSITIONS, [_row(0, 300), _row(300, 1620)]
        )
        assert [len(v) for v in pieces.values()] == [1, 1]
        assert skipped == []
        assert all(e["status"] == "full" for e in coverage)

    def test_row_crossing_a_gate_splits_at_the_gate(self):
        pieces, skipped, _ = bsc.assign_rows_to_edges(self.POSITIONS, [_row(0, 1620)])
        first = pieces[("M(0,0)", "M(0,1)")]
        second = pieces[("M(0,1)", "M(0,2)")]
        assert [(p["from_m"], p["to_m"]) for p in first] == [(0.0, 300.0)]
        assert [(p["from_m"], p["to_m"]) for p in second] == [(300.0, 1620.0)]
        # both pieces inherit the row's payload
        assert first[0]["row"]["depth_m"] == second[0]["row"]["depth_m"] == 1.6
        assert skipped == []

    def test_lengths_are_conserved_across_splits(self):
        pieces, _, _ = bsc.assign_rows_to_edges(self.POSITIONS, [_row(0, 1620)])
        total = sum(p["to_m"] - p["from_m"] for segs in pieces.values() for p in segs)
        assert total == 1620.0

    def test_row_beyond_the_last_gate_is_reported_not_emitted(self):
        pieces, skipped, coverage = bsc.assign_rows_to_edges(
            self.POSITIONS, [_row(0, 300), _row(300, 1620), _row(1620, 1820)]
        )
        assert sum(len(v) for v in pieces.values()) == 2
        assert [s["reason"] for s in skipped] == ["beyond_last_gate"]
        assert skipped[0]["from_m"] == 1620.0 and skipped[0]["to_m"] == 1820.0

    def test_row_without_cross_section_is_reported_not_defaulted(self):
        # The real case: LMC 0+000-0+170 is a Flume with no B/D/n/s survey.
        flume = _row(
            0,
            170,
            bottom_width_m=None,
            depth_m=None,
            manning_n=None,
            bed_slope=None,
            area_m2=None,
            note="Flume",
        )
        pieces, skipped, coverage = bsc.assign_rows_to_edges(
            self.POSITIONS, [flume, _row(170, 1620)]
        )
        assert [s["reason"] for s in skipped] == ["missing_cross_section"]
        [first_edge] = [e for e in coverage if e["downstream"] == "M(0,1)"]
        assert first_edge["status"] == "partial"
        assert first_edge["covered_m"] == 130.0
        assert first_edge["gap_m"] == 170.0
        assert first_edge["missing"] == [
            {
                "from_km": "0+000",
                "to_km": "0+170",
                "reason": "survey_row_missing_cross_section",
            }
        ]

    def test_unsurveyed_interior_and_trailing_chainage_is_measured(self):
        pieces, skipped, coverage = bsc.assign_rows_to_edges(
            self.POSITIONS, [_row(0, 300), _row(300, 800), _row(900, 1500)]
        )
        [edge2] = [e for e in coverage if e["downstream"] == "M(0,2)"]
        assert edge2["status"] == "partial"
        assert edge2["covered_m"] == 1100.0
        assert edge2["gap_m"] == 220.0
        assert edge2["missing"] == [
            {"from_km": "0+800", "to_km": "0+900", "reason": "unsurveyed"},
            {"from_km": "1+500", "to_km": "1+620", "reason": "unsurveyed"},
        ]

    def test_edge_with_no_rows_at_all_is_status_none(self):
        _, _, coverage = bsc.assign_rows_to_edges(self.POSITIONS, [_row(0, 300)])
        [edge2] = [e for e in coverage if e["downstream"] == "M(0,2)"]
        assert edge2["status"] == "none"
        assert edge2["covered_m"] == 0.0 and edge2["gap_m"] == 1320.0

    def test_non_increasing_gate_positions_fail_closed(self):
        with pytest.raises(bsc.WorkbookError):
            bsc.assign_rows_to_edges(
                [("M(0,0)", 0.0), ("M(0,1)", 300.0), ("M(0,2)", 300.0)],
                [_row(0, 300)],
            )

    def test_row_missing_design_area_or_discharge_is_reported_not_emitted(self):
        # A stale-formula workbook (saved without recalculation) presents as
        # blank A/Qd cells — emitting without slope/q_max would make consumers
        # default divergently (0.0 vs 1.0) and silently lose the capacity bound.
        no_area = _row(0, 300, area_m2=None)
        no_qd = _row(300, 1620, qd=None)
        pieces, skipped, _ = bsc.assign_rows_to_edges(self.POSITIONS, [no_area, no_qd])
        assert pieces == {}
        assert [s["reason"] for s in skipped] == [
            "missing_design_values",
            "missing_design_values",
        ]


class TestExtractionValidation:
    def test_gate_ids_that_collide_after_normalization_fail_closed(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for column, header in bsc.SHEET1_HEADERS.items():
            sheet.cell(row=2, column=column, value=header)
        for row_number, gate_id, chainage in (
            (3, "M(0,0)", "0+000"),
            (4, "M (0,0)", "0+300"),
        ):
            sheet.cell(row=row_number, column=2, value="Outlet")
            sheet.cell(row=row_number, column=3, value=row_number - 2)
            sheet.cell(row=row_number, column=5, value=gate_id)
            sheet.cell(row=row_number, column=14, value=1)
            sheet.cell(row=row_number, column=15, value="PC")
            sheet.cell(row=row_number, column=16, value=chainage)
            sheet.cell(row=row_number, column=21, value=1.0)

        with pytest.raises(bsc.WorkbookError, match="collide"):
            bsc.extract_gates(sheet)

    def test_overlapping_survey_rows_fail_closed(self):
        with pytest.raises(bsc.WorkbookError):
            bsc.validate_survey_rows([_row(0, 300), _row(200, 500)])

    def test_length_must_equal_chainage_delta(self):
        # The 1000x bug class: a length column that stops matching its own
        # chainage span must halt generation, not emit one of the two.
        with pytest.raises(bsc.WorkbookError):
            bsc.validate_survey_rows([_row(0, 300, length_m=3000.0)])

    def test_reversed_span_fails_closed(self):
        with pytest.raises(bsc.WorkbookError):
            bsc.validate_survey_rows([_row(300, 0, length_m=300.0)])

    def test_numeric_cell_holding_text_fails_closed(self):
        # Excel type drift: a q_max/Qd cell stored as the TEXT '9.961' must halt
        # generation, not silently become null and drop the capacity bound.
        with pytest.raises(bsc.WorkbookError):
            bsc.cell_number("9.961", "Sheet1 row 5 q_max")
        assert bsc.cell_number(None, "x") is None
        assert bsc.cell_number("-", "x") is None
        assert bsc.cell_number(" ", "x") is None
        assert bsc.cell_number(9.961, "x") == 9.961

    @pytest.mark.parametrize(
        "raw,expected",
        [("+198.504", 198.504), ("0.815", 0.815), (197.258, 197.258)],
    )
    def test_structure_numeric_text_is_normalized(self, raw, expected):
        assert bsc.structure_number(raw, "Sheet1 structure value") == expected

    @pytest.mark.parametrize("raw", [True, "198.5 m", "1e3", "--1", float("inf")])
    def test_malformed_structure_numeric_fails_closed(self, raw):
        with pytest.raises(bsc.WorkbookError):
            bsc.structure_number(raw, "Sheet1 structure value")

    def test_separator_row_with_stray_hydraulic_cells_fails_closed(self, tmp_path):
        # A row whose chainage cells were cleared but that still carries ANY
        # hydraulic value (e.g. only manning n / bed slope) is survey data being
        # silently dropped — not a separator.
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Characteristics"
        for col, text in bsc.CHAR_HEADERS_R2.items():
            ws.cell(row=2, column=col, value=text)
        for col, text in bsc.CHAR_HEADERS_R3.items():
            ws.cell(row=3, column=col, value=text)
        ws.cell(row=4, column=2, value="LMC")
        ws.cell(row=4, column=13, value=0.018)  # manning n, no chainage span
        path = tmp_path / "drift.xlsx"
        wb.save(path)
        loaded = openpyxl.load_workbook(path, data_only=True, read_only=True)
        with pytest.raises(bsc.WorkbookError, match="without a chainage span"):
            bsc.extract_survey_rows(loaded["Characteristics"])


class TestBuildGateCalibrations:
    @pytest.mark.parametrize(
        "k1,k2,r2",
        [(1.1, None, 0.99), (None, -1.2, 0.99), (1.1, -1.2, None)],
    )
    def test_rejects_partial_measured_calibration_triplets(self, k1, k2, r2):
        gate = {
            "gate_id": "M(0,0)",
            "k1": k1,
            "k2": k2,
            "r2": r2,
            "q_max": 11.2,
            "zone": 1,
            "width_m": None,
            "height_m": None,
        }
        with pytest.raises(bsc.WorkbookError, match="k1/k2/r2"):
            bsc.build_gate_calibrations([gate], Path("source.xlsx"), "source-sha")

    @pytest.mark.parametrize(
        "k1,k2,r2", [(0.0, -1.2, 0.99), (1.1, 0.0, 0.99), (1.1, -1.2, 1.01)]
    )
    def test_rejects_invalid_measured_calibration_values(self, k1, k2, r2):
        gate = {
            "gate_id": "M(0,0)",
            "k1": k1,
            "k2": k2,
            "r2": r2,
            "q_max": 11.2,
            "zone": 1,
            "width_m": None,
            "height_m": None,
        }
        with pytest.raises(bsc.WorkbookError, match="invalid measured k1/k2/r2"):
            bsc.build_gate_calibrations([gate], Path("source.xlsx"), "source-sha")


class TestSimilarGateInference:
    TARGET = {
        "gate_id": "TARGET",
        "canal_class": "SC",
        "shape": "rectangular",
        "width_m": 1.0,
        "height_m": 1.0,
    }
    DONORS = [
        {
            "gate_id": "A",
            "canal_class": "SC",
            "shape": "rectangular",
            "width_m": 2.0,
            "height_m": 2.0,
            "k1": 1.0,
            "k2": -1.0,
            "r2": 1.0,
        },
        {
            "gate_id": "B",
            "canal_class": "PC",
            "shape": "rectangular",
            "width_m": 1.0,
            "height_m": 1.0,
            "k1": 2.0,
            "k2": -3.0,
            "r2": 0.5,
        },
    ]

    def test_similarity_prioritizes_shape_size_and_canal_class(self):
        assert bsc.calibration_similarity(self.TARGET, self.DONORS[0]) == 0.825
        assert bsc.calibration_similarity(self.TARGET, self.DONORS[1]) == 0.8

    def test_inference_is_confidence_weighted_and_provenanced(self):
        assert bsc.infer_calibration(
            self.TARGET, self.DONORS, "similar-gate-v1:source-sha"
        ) == {
            "calibration_method": "inferred",
            "k1": 1.326531,
            "k2": -1.653061,
            "confidence": 0.435346,
            "source_gate_ids": ["A", "B"],
            "source_version": "similar-gate-v1:source-sha",
        }

    def test_inference_is_invariant_to_donor_input_order(self):
        assert bsc.infer_calibration(
            self.TARGET, list(reversed(self.DONORS)), "version"
        ) == bsc.infer_calibration(self.TARGET, self.DONORS, "version")

    def test_inference_fails_closed_without_a_similar_measured_gate(self):
        with pytest.raises(bsc.WorkbookError, match="measured donor"):
            bsc.infer_calibration(self.TARGET, [], "version")

    def test_zero_r2_donor_is_excluded_from_ranking_and_lineage(self):
        # 2.3-retro MEDIUM: an r2=0 donor carries zero confidence, so the loader
        # rejects any inference citing it (confidence can't sit below 0). The
        # generator must therefore drop r2<=0 donors before ranking/lineage —
        # here the highest-similarity donor (same shape/size/class) has r2=0.
        zero_r2_twin = {
            "gate_id": "TWIN",
            "canal_class": self.TARGET["canal_class"],
            "shape": self.TARGET["shape"],
            "width_m": self.TARGET["width_m"],
            "height_m": self.TARGET["height_m"],
            "k1": 9.0,
            "k2": -9.0,
            "r2": 0.0,
        }
        with_zero = bsc.infer_calibration(
            self.TARGET, [zero_r2_twin] + self.DONORS, "similar-gate-v1:source-sha"
        )
        assert "TWIN" not in with_zero["source_gate_ids"]
        # dropping a zero-weight donor leaves the surviving inference untouched
        assert with_zero == bsc.infer_calibration(
            self.TARGET, self.DONORS, "similar-gate-v1:source-sha"
        )

    def test_all_zero_r2_donors_fail_closed(self):
        zeros = [{**donor, "r2": 0.0} for donor in self.DONORS]
        with pytest.raises(bsc.WorkbookError, match="measured donor"):
            bsc.infer_calibration(self.TARGET, zeros, "version")

    def test_zero_r2_same_shape_donor_fails_closed_not_cross_shape(self):
        # QCHECK (workflow): dropping an r2<=0 donor must NOT let a circular target
        # silently borrow a differently-shaped donor's flow law. When the only
        # same-shape measured donor has r2<=0, fail closed exactly as the code did
        # before the r2 filter (candidates=same_shape -> zero weight -> WorkbookError),
        # rather than reaching for a rectangular donor's k1/k2 for a circular gate.
        circular_target = {
            "gate_id": "CIRC",
            "canal_class": "FTO",
            "shape": "circular",
            "width_m": 0.4,
            "height_m": 0.4,
        }
        zero_r2_circular = {
            "gate_id": "CDONOR",
            "canal_class": "FTO",
            "shape": "circular",
            "width_m": 0.4,
            "height_m": 0.4,
            "k1": 1.3,
            "k2": -3.0,
            "r2": 0.0,
        }
        valid_rectangular = {
            "gate_id": "RECT",
            "canal_class": "FTO",
            "shape": "rectangular",
            "width_m": 2.0,
            "height_m": 2.0,
            "k1": 1.2,
            "k2": -1.3,
            "r2": 0.9,
        }
        with pytest.raises(bsc.WorkbookError, match="measured donor"):
            bsc.infer_calibration(
                circular_target, [zero_r2_circular, valid_rectangular], "version"
            )

    def test_subnormal_r2_weight_underflow_fails_closed(self):
        # QCHECK LOW (gpt-5.6-sol): score>0 and r2>0 do not guarantee a positive
        # product — a normal similarity times a subnormal r2 can underflow to 0.0.
        # The generator must fail closed, not divide by zero. A far-off-dimension
        # donor scores ~0.20, and 0.20 * 5e-324 rounds to 0.0.
        target = {
            "gate_id": "T",
            "canal_class": "SC",
            "shape": None,
            "width_m": 1.0,
            "height_m": 1.0,
        }
        subnormal_donor = {
            "gate_id": "D",
            "canal_class": "SC",
            "shape": "rectangular",
            "width_m": 100.0,
            "height_m": 100.0,
            "k1": 1.0,
            "k2": -1.0,
            "r2": 5e-324,
        }
        with pytest.raises(bsc.WorkbookError, match="underflow"):
            bsc.infer_calibration(target, [subnormal_donor], "version")


@pytest.fixture(scope="module")
def artifacts():
    return bsc.build_all(WORKBOOK)


class TestRealWorkbookGeneration:
    def test_v3_source_path_and_hash_are_exact(self):
        assert WORKBOOK.is_file()
        assert hashlib.sha256(WORKBOOK.read_bytes()).hexdigest() == WORKBOOK_SHA256
        assert bsc.DEFAULT_WORKBOOK == WORKBOOK

    def test_all_artifacts_carry_the_workbook_sha256(self, artifacts):
        digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
        for name in (
            "network",
            "canal_geometry",
            "geometry_coverage",
            "gate_calibrations",
        ):
            meta = artifacts[name]["metadata"]
            assert meta["source_sha256"] == digest
            assert meta["source_workbook"] == WORKBOOK.name

    def test_gate_calibrations_preserve_measured_provenance(self, artifacts):
        digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
        calibrations = artifacts["gate_calibrations"]
        assert calibrations["metadata"]["gates_by_calibration_method"] == {
            "measured": 10,
            "inferred": 48,
            "default": 0,
        }
        assert calibrations["gates"]["M(0,0)"] == {
            "gate_id": "M(0,0)",
            "calibration_method": "measured",
            "k1": 1.0693,
            "k2": -1.229,
            "confidence": 0.9986,
            "source_gate_ids": ["M(0,0)"],
            "source_version": digest,
            "canal_class": "PC",
            "q_max_m3s": 11.2,
            "zone": 1,
            "design_fsl_msl_m": 221,
            "sill_msl_m": 204.5,
            "structure_max_flow_m3s": 11.2,
            "design_fsl_reference_side": "upstream",
            "structure_data_status": "complete",
            "structure_role": "control",
        }

    def test_all_generated_gate_ids_are_compact_and_collision_free(self, artifacts):
        from core.node_id import normalize_gate_id

        for artifact_name in ("network", "gate_calibrations"):
            gate_ids = list(artifacts[artifact_name]["gates"])
            assert len(gate_ids) == 58
            assert all(gate_id == normalize_gate_id(gate_id) for gate_id in gate_ids)

    def test_eight_complete_rmc_controls_carry_exact_v3_structure_fields(
        self, artifacts
    ):
        gates = artifacts["gate_calibrations"]["gates"]
        expected = {
            "M(0,0;2,0)": (205.561, 203.712, 1.2),
            "M(0,0;2,1)": (203.888, 202.938, 1.234),
            "M(0,0;2,2)": (198.504, 197.954, 1.234),
            "M(0,0;2,3)": (198.308, 197.258, 0.252),
            "M(0,0;2,1;1,0)": (200.204, 199.404, 0.815),
            "M(0,0;2,1;1,1)": (200.054, 199.254, 0.67),
            "M(0,0;2,1;1,2)": (198.084, 197.334, 0.397),
            "M(0,0;2,1;1,3)": (196.934, 196.284, 0.205),
        }
        assert {
            gate_id: (
                gates[gate_id]["design_fsl_msl_m"],
                gates[gate_id]["sill_msl_m"],
                gates[gate_id]["structure_max_flow_m3s"],
            )
            for gate_id in expected
        } == expected
        assert all(
            gates[gate_id]["design_fsl_reference_side"] == "upstream"
            and gates[gate_id]["structure_data_status"] == "complete"
            and gates[gate_id]["structure_role"] == "control"
            for gate_id in expected
        )

    def test_tail_and_turnout_structure_statuses_are_explicit(self, artifacts):
        gates = artifacts["gate_calibrations"]["gates"]
        assert {
            gate_id: (
                gates[gate_id]["structure_role"],
                gates[gate_id]["structure_data_status"],
                gates[gate_id]["design_fsl_msl_m"],
                gates[gate_id]["sill_msl_m"],
                gates[gate_id]["structure_max_flow_m3s"],
            )
            for gate_id in (
                "M(0,0;2,1;1,4)",
                "M(0,0;2,1;1,2;1,0)",
            )
        } == {
            "M(0,0;2,1;1,4)": (
                "tail",
                "incomplete",
                195.354,
                194.854,
                None,
            ),
            "M(0,0;2,1;1,2;1,0)": (
                "turnout",
                "unavailable",
                None,
                None,
                None,
            ),
        }

    def test_gate_calibrations_mark_unmeasured_rows_as_provisional_inferences(
        self, artifacts
    ):
        digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
        calibrations = artifacts["gate_calibrations"]
        measured = {
            gate_id
            for gate_id, gate in calibrations["gates"].items()
            if gate["calibration_method"] == "measured"
        }
        inferred = [
            gate
            for gate in calibrations["gates"].values()
            if gate["calibration_method"] == "inferred"
        ]
        measured_records = [
            gate
            for gate in calibrations["gates"].values()
            if gate["calibration_method"] == "measured"
        ]
        assert calibrations["metadata"]["intended_use"] == "planning_only"
        assert len(inferred) == 48
        assert all(0.0 < gate["confidence"] < 0.9805 for gate in inferred)
        assert all(1 <= len(gate["source_gate_ids"]) <= 3 for gate in inferred)
        assert all(set(gate["source_gate_ids"]) <= measured for gate in inferred)
        assert all(
            gate["source_version"] == f"similar-gate-v1:{digest}" for gate in inferred
        )
        assert all(
            min(gate["k1"] for gate in measured_records)
            <= gate["k1"]
            <= max(gate["k1"] for gate in measured_records)
            and min(gate["k2"] for gate in measured_records)
            <= gate["k2"]
            <= max(gate["k2"] for gate in measured_records)
            for gate in inferred
        )

    def test_gate_ratings_match_the_network_artifact(self, artifacts):
        calibration_gates = artifacts["gate_calibrations"]["gates"]
        network_gates = artifacts["network"]["gates"]
        assert {
            gate_id: gate.get("q_max_m3s")
            for gate_id, gate in calibration_gates.items()
        } == {gate_id: gate.get("q_max") for gate_id, gate in network_gates.items()}

    def test_circular_gate_height_is_used_as_its_hydraulic_width(self, artifacts):
        gate = artifacts["gate_calibrations"]["gates"]["M(0,0;2,0;1,0)"]
        assert (gate["shape"], gate["width_m"], gate["height_m"]) == (
            "circular",
            0.4,
            0.4,
        )

    def test_geometry_passes_the_strict_runtime_loader(self, artifacts, tmp_path):
        path = tmp_path / "canal_geometry.json"
        path.write_text(json.dumps(artifacts["canal_geometry"]), encoding="utf-8")
        data = load_canal_geometry_config(str(path))
        # 99 survey rows - 1 flume (no cross-section) - 8 post-terminal tails
        # + 4 split-at-gate pieces = 94 emitted segments.
        assert data["summary"]["total_sections"] == len(data["canal_sections"]) == 94

    def test_network_passes_the_strict_runtime_loader(self, artifacts, tmp_path):
        path = tmp_path / "network.json"
        path.write_text(json.dumps(artifacts["network"]), encoding="utf-8")
        data = load_network_config(str(path))
        assert len(data["gates"]) == 58 and len(data["edges"]) == 58

    def test_network_edges_equal_the_naming_grammar_derivation(self, artifacts):
        net = artifacts["network"]
        assert net["edges"] == [list(e) for e in edges_from_names(list(net["gates"]))]

    def test_geometry_feeds_the_multisegment_runtime_indexer(self, artifacts):
        sections = sections_by_edge_from_geometry(artifacts["canal_geometry"])
        assert len(sections) == 41  # every serial gate-to-gate reach is surveyed

    def test_first_lmc_reach_is_the_post_flume_concrete_piece(self, artifacts):
        # M(0,1) does not exist. The first LMC reach runs directly from M(0,0)
        # to M(0,2); its flume [0,170] has no cross-section.
        sections = sections_by_edge_from_geometry(artifacts["canal_geometry"])
        [seg] = sections[("M(0,0)", "M(0,2)")]
        assert seg["length_m"] == 1450
        assert seg["from_km_m"] == 170.0 and seg["to_km_m"] == 1620.0
        assert seg["q_max"] == 9.961  # Qd of the source row
        assert seg["seepage_rate_m_s"] == 1.0e-5  # concrete

    def test_lmc_tail_reach_stops_before_the_post_terminal_earth_row(self, artifacts):
        sections = sections_by_edge_from_geometry(artifacts["canal_geometry"])
        segs = sections[("M(0,13)", "M(0,14)")]
        assert [s["length_m"] for s in segs] == [240, 160, 200]
        assert segs[-1]["seepage_rate_m_s"] == 1.0e-5

    def test_side_slope_is_derived_not_the_lining_thickness(self, artifacts):
        slopes = {
            s["geometry"]["cross_section"]["side_slope"]
            for s in artifacts["canal_geometry"]["canal_sections"]
        }
        assert slopes == {1.5}

    def test_zone_summary_is_consistent(self, artifacts):
        sections = artifacts["canal_geometry"]["canal_sections"]
        by_zone = {}
        for s in sections:
            by_zone[f"zone_{s['zone']}"] = by_zone.get(f"zone_{s['zone']}", 0) + 1
        assert artifacts["canal_geometry"]["summary"]["by_zone"] == by_zone

    def test_coverage_report_covers_all_58_edges(self, artifacts):
        edges = artifacts["geometry_coverage"]["edges"]
        assert len(edges) == 58
        by_status = {}
        for e in edges:
            by_status[e["status"]] = by_status.get(e["status"], 0) + 1
        assert by_status == {"full": 40, "partial": 1, "not_applicable": 17}
        by_category = {}
        for e in edges:
            by_category[e["category"]] = by_category.get(e["category"], 0) + 1
        assert by_category == {
            "serial": 41,
            "junction_head": 13,
            "offtake": 3,
            "source": 1,
        }

    def test_partial_reaches_and_gaps_match_the_survey(self, artifacts):
        gaps = {
            (e["upstream"], e["downstream"]): e["gap_m"]
            for e in artifacts["geometry_coverage"]["edges"]
            if e["status"] == "partial"
        }
        assert gaps == {
            ("M(0,0)", "M(0,2)"): 170.0,
        }

    def test_skipped_rows_are_reported(self, artifacts):
        skipped = artifacts["geometry_coverage"]["summary"]["skipped_rows"]
        assert skipped == {"missing_cross_section": 1, "beyond_last_gate": 8}

    def test_structure_fields_are_scoped_to_the_planning_only_bundle(self, artifacts):
        for name in ("network", "canal_geometry", "geometry_coverage"):
            assert "sill_msl_m" not in json.dumps(artifacts[name])
        assert artifacts["gate_calibrations"]["metadata"]["intended_use"] == (
            "planning_only"
        )

    def test_reaches_block_carries_every_serial_span(self, artifacts):
        # The runtime needs the gate-to-gate span to measure head/tail survey
        # gaps (reach_chainage_gap_m) and physical reach length (legacy solver).
        reaches = artifacts["canal_geometry"]["reaches"]
        assert len(reaches) == 41
        by_edge = {(r["from_node"], r["to_node"]): r for r in reaches}
        first = by_edge[("M(0,0)", "M(0,2)")]
        assert (first["from_km"], first["to_km"]) == ("0+000", "1+620")
        assert first["span_m"] == 1620
        assert first["covered_m"] == 1450 and first["gap_m"] == 170
        assert all(r["span_m"] == r["covered_m"] + r["gap_m"] for r in reaches)

    def test_rotation_plan_q_is_retained_alongside_qd(self, artifacts):
        # Qd is the binding q_max; the rotation-plan Qmax column is preserved as
        # q_rotation_plan so no committed capacity information is deleted
        # (capacity governance moves to the 2.3 calibration schema).
        sections = artifacts["canal_geometry"]["canal_sections"]
        first = sections[0]["geometry"]["hydraulic_params"]
        assert first["q_max"] == 9.961
        assert first["q_rotation_plan"] == 9.926
        rmc = [s for s in sections if s["canal_name"] == "RMC"]
        assert rmc and all(
            "q_rotation_plan" not in s["geometry"]["hydraulic_params"] for s in rmc
        )

    def test_every_v3_gate_has_a_numeric_zone(self, artifacts):
        assert all(
            isinstance(gate["zone"], int)
            for gate in artifacts["network"]["gates"].values()
        )

    def test_every_section_carries_slope_and_q_max(self, artifacts):
        # area_m2 and qd are emission-required, so no section can ship without
        # its slope or capacity bound — consumers never fall back to divergent
        # defaults (conveyance uses 0.0, the legacy solver 1.0).
        for section in artifacts["canal_geometry"]["canal_sections"]:
            assert "side_slope" in section["geometry"]["cross_section"]
            assert "q_max" in section["geometry"]["hydraulic_params"]


class TestCoverageClassifier:
    GATES = {
        "M(0,0)": {"canal": "Outlet"},
        "M(0,1)": {"canal": "LMC"},
        "M(0,0;1,0)": {"canal": "WW"},
        "M(0,1;1,0)": {"canal": "New"},
        "M(0,1;1,1)": {"canal": "New"},
    }

    def test_unmatched_multi_gate_chain_edges_are_serial_none_not_offtake(self):
        # QCHECK 2.1b HIGH: if a whole canal's survey block disappears from the
        # workbook, its serial reaches must show as status 'none' — labelling
        # them 'offtake'/'not_applicable' certifies the loss as normal.
        entries = bsc.classify_edges(
            gate_ids=list(self.GATES),
            canal_by_gate={k: v["canal"] for k, v in self.GATES.items()},
            km_by_gate={
                "M(0,0)": "0+000",
                "M(0,1)": "0+300",
                "M(0,0;1,0)": "0+160",
                "M(0,1;1,0)": "0+000",
                "M(0,1;1,1)": "2+600",
            },
            matched_head_canals={"M(0,0)": "LMC"},
            serial_coverage={
                ("M(0,0)", "M(0,1)"): {
                    "span_m": 300,
                    "covered_m": 300,
                    "gap_m": 0,
                    "segments": 1,
                    "status": "full",
                    "missing": [],
                }
            },
        )
        by_edge = {(e["upstream"], e["downstream"]): e for e in entries}
        assert by_edge[("S", "M(0,0)")]["category"] == "source"
        assert by_edge[("M(0,0)", "M(0,1)")]["category"] == "serial"
        # single-gate unmatched chain -> offtake (no canal of its own to survey)
        assert by_edge[("M(0,0)", "M(0,0;1,0)")]["category"] == "offtake"
        # unmatched MULTI-gate chain: head is a junction, serial edge is a
        # visible survey hole
        assert by_edge[("M(0,1)", "M(0,1;1,0)")]["category"] == "junction_head"
        serial_none = by_edge[("M(0,1;1,0)", "M(0,1;1,1)")]
        assert serial_none["category"] == "serial"
        assert serial_none["status"] == "none"
        assert serial_none["missing"] == [{"reason": "canal_not_in_survey"}]


class TestRegenerationLock:
    """The committed artifacts must be BYTE-exactly what the committed workbook
    regenerates — hand edits, reformatting, and stale artifacts fail the suite."""

    @pytest.mark.parametrize(
        "name,filename",
        [
            ("network", "network.json"),
            ("canal_geometry", "canal_geometry.json"),
            ("geometry_coverage", "geometry_coverage.json"),
            ("gate_calibrations", "gate_calibrations.json"),
        ],
    )
    def test_committed_artifact_equals_regeneration(self, artifacts, name, filename):
        expected = (
            json.dumps(artifacts[name], indent=2, ensure_ascii=False) + "\n"
        ).encode("utf-8")
        assert (CONFIG_DIR / filename).read_bytes() == expected


class TestMainCli:
    def test_writes_four_deterministic_artifacts(self, tmp_path):
        out1, out2 = tmp_path / "a", tmp_path / "b"
        bsc.main([str(WORKBOOK), "--out-dir", str(out1)])
        bsc.main([str(WORKBOOK), "--out-dir", str(out2)])
        names = [
            "network.json",
            "canal_geometry.json",
            "geometry_coverage.json",
            "gate_calibrations.json",
        ]
        for n in names:
            b1, b2 = (out1 / n).read_bytes(), (out2 / n).read_bytes()
            assert b1 == b2  # deterministic: no timestamps, stable ordering
            assert b1.endswith(b"\n")
