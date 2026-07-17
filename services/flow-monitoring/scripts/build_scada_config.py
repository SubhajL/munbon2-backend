#!/usr/bin/env python3
"""
Wave 2.1b — strict SCADA workbook -> canonical config generator.

Replaces excel_to_canal_sections.py, whose Sheet1 `ระยะทาง (เมตร)` column holds
KILOMETRES and was written straight into `length_m` (the 1000x bug), and whose
missing survey columns silently became invented defaults (10 m widths, 0.5 m
freeboards, fabricated sills).

Sources (workbook `SCADA Section Detailed Information 2026-07-14 V3.0 SL.xlsx`):
- Sheet1 — the gates: exact network ids, canal, zone, per-gate chainage
  (`km`), rated q_max, service area. The `km` column is the gate-position
  authority; the ระยะทาง column is validated against consecutive existing
  valves without renumbering sparse physical SCADA ids.
- Characteristics — 99 surveyed hydraulic segments keyed canal + chainage span
  (NOT gate ids), lengths in true metres (validated == chainage delta). The
  right-hand water-volume table (col >= 28) is ignored.
- Sheet1 AH-AJ — approved design FSL, structure sill, and structure maximum
  flow. FSL is the upstream-of-structure design reference, not a live level;
  structure sill is never treated as the canal bed.

Semantics:
- A survey row crossing a gate position is split at the gate; every emitted
  piece is pure chainage arithmetic on real data (never prorated estimates —
  row length == chainage delta is validated first).
- Segment q_max is the design discharge `Qd` (present on every hydraulic row);
  `Qmax (จากแผนรอบเวรส่งน้ำ)` is a rotation-plan value, absent for whole canals.
- side_slope is derived from the design flow area: A = D*(B + z*D) so
  z = (A/D - B)/D. The old file's 0.06 was the `t` column — the 6 cm lining
  THICKNESS misread as a slope. The workbook yields z = 1.5 on all rows.
- Anything unmappable (a row with no cross-section, chainage beyond the last
  gate, an unsurveyed reach) is measured and REPORTED in geometry_coverage.json
  — never defaulted.

Artifacts (all stamped with the workbook SHA-256):
  src/config/network.json            gates + edges_from_names topology
  src/config/canal_geometry.json     2.1a multi-segment survey geometry
  src/config/geometry_coverage.json  per-edge coverage report
  src/config/gate_calibrations.json  measured/inferred calibration provenance

Run:  python scripts/build_scada_config.py [workbook.xlsx] [--out-dir DIR]
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from pathlib import Path

SERVICE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SERVICE_ROOT / "src"))

from core.conveyance_loss import parse_chainage_m  # noqa: E402
from core.network_topology import ROOT, edges_from_names  # noqa: E402
from core.node_id import normalize_gate_id, parse_gate_id  # noqa: E402
from core.routing_topology import (  # noqa: E402
    derive_routing_topology,
    routing_topology_payload,
)

DEFAULT_WORKBOOK = (
    SERVICE_ROOT
    / "data"
    / "sources"
    / "scada"
    / "SCADA Section Detailed Information 2026-07-14 V3.0 SL.xlsx"
)
DEFAULT_OUT_DIR = SERVICE_ROOT / "src" / "config"
GENERATOR = "scripts/build_scada_config.py"

# Consumed columns, validated against the header rows before any cell is read —
# a silently shifted column is exactly how the 1000x bug happened.
SHEET1_HEADERS = {
    2: "Canal Name",
    5: "Gate Valve",
    14: "Zone",
    15: "PC/SC/TC/QC",
    16: "km",
    17: "Area (Rais)",
    20: "ระยะทาง (เมตร)",
    21: "q_max (m^3/s)",
    22: "k1",
    23: "k2",
    24: "r2",
    25: "width (m)",
    26: "height (m)",
    34: "FSL (m-MSL)",
    35: "Sill Level (m-MSL)",
    36: "Max Flow (m^3/s)",
}
CHAR_HEADERS_R2 = {
    2: "คลอง",
    5: "ระยะ",
    7: "Qmax (จากแผนรอบเวรส่งน้ำ)",
    8: "Qd",
    9: "A",
    13: "n",
    14: "s",
    15: "B",
    16: "D",
    26: "หมายเหตุ",
}
CHAR_HEADERS_R3 = {3: "เริ่ม", 4: "สิ้นสุด"}
# Columns 6..25 hold the LEFT table's hydraulic values; a chainage-less row with
# any of them populated is dropped survey data, not a separator.
CHAR_HYDRAULIC_COLS = range(6, 26)

LENGTH_TOLERANCE_M = 0.5
MAX_SIDE_SLOPE = 5.0
CHAINAGE_EPS = 1e-6
MAX_SIMILAR_GATES = 3
INFERENCE_CONFIDENCE_PENALTY = 0.7
SHAPE_SIMILARITY_WEIGHT = 0.45
DIMENSION_SIMILARITY_WEIGHT = 0.35
CANAL_CLASS_SIMILARITY_WEIGHT = 0.2
DESIGN_FSL_REFERENCE_SIDE = "upstream"
_STRUCTURE_NUMBER = re.compile(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)\Z")


class WorkbookError(ValueError):
    """The workbook violates an assumption this generator depends on."""


def parse_km_marker(value) -> float:
    """Chainage ``K+MMM`` (e.g. ``6+880``) in metres; unparseable input raises.

    The grammar is core.conveyance_loss.parse_chainage_m's — ONE chainage
    grammar for generator and runtime. Only the failure mode differs: the
    runtime decides whether ordering is required, generation fails closed
    (an unreadable marker would silently unmap a survey row).
    """
    metres = parse_chainage_m(value)
    if metres is None:
        raise WorkbookError(f"unparseable chainage marker {value!r}")
    return metres


def format_km_marker(metres: float) -> str:
    """Metres as a zero-padded ``K+MMM`` marker; round-trips parse_km_marker."""
    km, rem = int(metres // 1000.0), metres - int(metres // 1000.0) * 1000.0
    if rem == int(rem):
        return f"{km}+{int(rem):03d}"
    return f"{km}+{rem:07.3f}".rstrip("0").rstrip(".")


def derive_side_slope(area_m2: float, bottom_width_m: float, depth_m: float) -> float:
    """Trapezoid side slope from the design flow area: A = D*(B + z*D).

    The survey has no explicit slope column (`t` is the lining thickness); the
    area column determines z exactly. Out-of-range results mean inconsistent
    survey values — fail closed rather than emit bad geometry.
    """
    if not (bottom_width_m > 0 and depth_m > 0 and area_m2 > 0):
        raise WorkbookError(
            f"cross-section values must be positive, got A={area_m2!r}"
            f" B={bottom_width_m!r} D={depth_m!r}"
        )
    z = round((area_m2 / depth_m - bottom_width_m) / depth_m, 2)
    if z == 0.0:
        z = 0.0  # normalize -0.0
    if not (0.0 <= z <= MAX_SIDE_SLOPE):
        raise WorkbookError(
            f"derived side slope {z} outside [0, {MAX_SIDE_SLOPE}] for"
            f" A={area_m2} B={bottom_width_m} D={depth_m}"
        )
    return z


def lining_from_note(note) -> str | None:
    """Lining from the หมายเหตุ column; None when the note states no lining.

    'คลองดาดคอนกรีต' (and the คาด typo) -> concrete; 'คลองดิน' -> earth. Other
    remarks ('Flume', 'Hl=3.00-6.00', blank) carry no lining statement — the
    runtime treats an absent lining as the documented 'unknown' seepage rate.
    """
    if not isinstance(note, str):
        return None
    if "คอนกรีต" in note:
        return "concrete"
    if "ดิน" in note:
        return "earth"
    return None


def cell_number(value, where: str):
    """A finite number from a numeric cell; blanks and '-' are None.

    ANY other text raises: a numeric column drifting to Excel text format
    (a q_max stored as the TEXT '9.961') must halt generation, not silently
    become null and drop a capacity bound."""
    if value is None:
        return None
    if isinstance(value, bool):
        raise WorkbookError(f"{where}: boolean cell in a numeric column")
    if isinstance(value, (int, float)):
        return float(value) if value == value else None  # NaN cell = blank
    if isinstance(value, str) and value.strip() in ("", "-"):
        return None
    raise WorkbookError(f"{where}: non-numeric cell {value!r} in a numeric column")


def structure_number(value, where: str):
    """Finite AH-AJ number, accepting the workbook's decimal-text cells."""
    if value is None:
        return None
    if isinstance(value, bool):
        raise WorkbookError(f"{where}: boolean cell in a structure numeric column")
    if isinstance(value, (int, float)):
        number = float(value)
    elif isinstance(value, str):
        text = value.strip()
        if text in ("", "-"):
            return None
        if _STRUCTURE_NUMBER.fullmatch(text) is None:
            raise WorkbookError(f"{where}: malformed structure numeric cell {value!r}")
        number = float(text)
    else:
        raise WorkbookError(f"{where}: invalid structure numeric cell {value!r}")
    if not math.isfinite(number):
        raise WorkbookError(f"{where}: non-finite structure numeric cell {value!r}")
    return number


def classify_structure_role(
    gate_id: str, canal_class: str, section_number: float | None
) -> str:
    if canal_class == "FTO":
        return "turnout"
    if section_number is None and canal_class in {"PC", "SC", "TC", "QC"}:
        return "tail"
    if canal_class in {"PC", "SC", "TC", "QC", "WW"}:
        return "control"
    return "unknown"


def _intish(value: float):
    """Integral floats as ints, so emitted JSON matches the survey's integers."""
    return int(value) if float(value).is_integer() else value


def _validate_headers(ws, row: int, expected: dict, sheet: str) -> None:
    cells = {c: ws.cell(row=row, column=c).value for c in expected}
    for col, want in expected.items():
        got = str(cells[col]).strip() if cells[col] is not None else None
        if got != want:
            raise WorkbookError(
                f"{sheet} header drift at row {row} col {col}:"
                f" expected {want!r}, got {got!r}"
            )


def extract_gates(ws) -> list[dict]:
    """Sheet1 gate rows: exact ids, geometry, ratings and calibration triplets."""
    _validate_headers(ws, 2, SHEET1_HEADERS, "Sheet1")
    gates = []
    seen_gate_ids = set()
    for row in ws.iter_rows(min_row=3, max_col=36):
        cell = lambda col: row[col - 1].value  # noqa: E731
        raw_id = cell(5)
        if raw_id is None:
            continue
        gate_id = normalize_gate_id(str(raw_id).strip())
        if gate_id in seen_gate_ids:
            raise WorkbookError(
                f"Sheet1 gate ids collide after normalization at {gate_id}"
            )
        seen_gate_ids.add(gate_id)
        parsed = parse_gate_id(gate_id)  # NodeIdError propagates: bad id = bad workbook
        canal = str(cell(2)).strip() if cell(2) is not None else None
        section_number = cell_number(cell(3), f"Sheet1 gate {gate_id} Section")
        zone = cell_number(cell(14), f"Sheet1 gate {gate_id} Zone")
        canal_class = str(cell(15)).strip() if cell(15) is not None else None
        if not canal_class:
            raise WorkbookError(f"gate {gate_id}: missing canal class")
        q_max = cell_number(cell(21), f"Sheet1 gate {gate_id} q_max")
        if q_max is not None and q_max <= 0:
            raise WorkbookError(f"gate {gate_id}: q_max must be > 0, got {q_max}")
        raw_width = cell(25)
        if isinstance(raw_width, str) and raw_width.strip().upper() == "C":
            shape = "circular"
            width_m = None
        else:
            width_m = cell_number(raw_width, f"Sheet1 gate {gate_id} width")
            shape = "rectangular" if width_m is not None else None
        height_m = cell_number(cell(26), f"Sheet1 gate {gate_id} height")
        for name, value in (("width", width_m), ("height", height_m)):
            if value is not None and value <= 0:
                raise WorkbookError(f"gate {gate_id}: {name} must be > 0, got {value}")
        if shape == "circular":
            if height_m is None:
                raise WorkbookError(
                    f"gate {gate_id}: circular gate requires a numeric diameter"
                )
            width_m = height_m
        design_fsl_msl_m = structure_number(
            cell(34), f"Sheet1 gate {gate_id} design FSL"
        )
        sill_msl_m = structure_number(cell(35), f"Sheet1 gate {gate_id} structure sill")
        structure_max_flow_m3s = structure_number(
            cell(36), f"Sheet1 gate {gate_id} structure max flow"
        )
        if structure_max_flow_m3s is not None and structure_max_flow_m3s <= 0:
            raise WorkbookError(
                f"gate {gate_id}: structure max flow must be > 0,"
                f" got {structure_max_flow_m3s}"
            )
        if (
            design_fsl_msl_m is not None
            and sill_msl_m is not None
            and design_fsl_msl_m <= sill_msl_m
        ):
            raise WorkbookError(
                f"gate {gate_id}: design FSL must exceed structure sill"
            )
        structure_presence = (
            design_fsl_msl_m is not None,
            sill_msl_m is not None,
            structure_max_flow_m3s is not None,
        )
        structure_data_status = (
            "complete"
            if all(structure_presence)
            else "unavailable"
            if not any(structure_presence)
            else "incomplete"
        )
        gates.append(
            {
                "gate_id": gate_id,
                "canal": canal,
                "zone": int(zone) if zone is not None else None,
                "canal_class": canal_class,
                "parsed": [list(t) for t in parsed],
                "km_text": str(cell(16)).strip() if cell(16) is not None else None,
                "km_m": parse_km_marker(cell(16)),
                "distance_km": cell_number(cell(20), f"Sheet1 gate {gate_id} distance"),
                "q_max": q_max,
                "area": cell_number(cell(17), f"Sheet1 gate {gate_id} Area"),
                "k1": cell_number(cell(22), f"Sheet1 gate {gate_id} k1"),
                "k2": cell_number(cell(23), f"Sheet1 gate {gate_id} k2"),
                "r2": cell_number(cell(24), f"Sheet1 gate {gate_id} r2"),
                "shape": shape,
                "width_m": width_m,
                "height_m": height_m,
                "design_fsl_msl_m": design_fsl_msl_m,
                "sill_msl_m": sill_msl_m,
                "structure_max_flow_m3s": structure_max_flow_m3s,
                "design_fsl_reference_side": (
                    DESIGN_FSL_REFERENCE_SIDE if design_fsl_msl_m is not None else None
                ),
                "structure_data_status": structure_data_status,
                "structure_role": classify_structure_role(
                    gate_id, canal_class, section_number
                ),
            }
        )
    return gates


def extract_survey_rows(ws) -> list[dict]:
    """Characteristics LEFT-table rows (canal forward-filled, chainage strict).

    The right-hand water-volume table (col >= 28) is never read. Separator rows
    must be empty across the consumed columns — a half-filled row is drift.
    """
    _validate_headers(ws, 2, CHAR_HEADERS_R2, "Characteristics")
    _validate_headers(ws, 3, CHAR_HEADERS_R3, "Characteristics")
    rows = []
    canal = None
    for rn, row in enumerate(ws.iter_rows(min_row=4, max_col=26), start=4):
        cell = lambda col: row[col - 1].value  # noqa: E731
        if cell(2) is not None and str(cell(2)).strip():
            canal = " ".join(str(cell(2)).split())
        where = f"Characteristics row {rn}"
        from_raw, to_raw, length = cell(3), cell(4), cell_number(cell(5), where)
        if from_raw is None and to_raw is None and length is None:
            stray = [
                c
                for c in CHAR_HYDRAULIC_COLS
                if cell(c) is not None and str(cell(c)).strip() not in ("", "-")
            ]
            if stray:
                raise WorkbookError(
                    f"{where}: hydraulic data without a chainage span"
                    f" (columns {stray})"
                )
            continue
        if canal is None:
            raise WorkbookError(f"{where}: data before any canal name")
        from_m, to_m = parse_km_marker(from_raw), parse_km_marker(to_raw)
        if length is None:
            raise WorkbookError(f"{where}: missing length")
        rows.append(
            {
                "canal": canal,
                "from_m": from_m,
                "to_m": to_m,
                "length_m": length,
                "q_rotation": cell_number(cell(7), f"{where} Qmax"),
                "qd": cell_number(cell(8), f"{where} Qd"),
                "area_m2": cell_number(cell(9), f"{where} A"),
                "manning_n": cell_number(cell(13), f"{where} n"),
                "bed_slope": cell_number(cell(14), f"{where} s"),
                "bottom_width_m": cell_number(cell(15), f"{where} B"),
                "depth_m": cell_number(cell(16), f"{where} D"),
                "note": str(cell(26)).strip() if cell(26) is not None else None,
                "excel_row": rn,
            }
        )
    return rows


def validate_survey_rows(rows: list[dict]) -> None:
    """Per canal: spans run forward, length == chainage delta, no overlaps."""
    by_canal: dict = {}
    for row in rows:
        by_canal.setdefault(row["canal"], []).append(row)
    for canal, group in by_canal.items():
        for row in group:
            span = row["to_m"] - row["from_m"]
            if span <= 0:
                raise WorkbookError(
                    f"{canal} row {row['excel_row']}: span must run forward,"
                    f" got {row['from_m']}..{row['to_m']}"
                )
            if abs(span - row["length_m"]) > LENGTH_TOLERANCE_M:
                raise WorkbookError(
                    f"{canal} row {row['excel_row']}: length {row['length_m']} m"
                    f" does not match chainage span {span} m — the length-column"
                    " unit drifted (the excel_to_canal_sections 1000x bug class)"
                )
        ordered = sorted(group, key=lambda r: r["from_m"])
        for prev, nxt in zip(ordered, ordered[1:]):
            if nxt["from_m"] < prev["to_m"] - CHAINAGE_EPS:
                raise WorkbookError(
                    f"{canal}: survey rows overlap at chainage {nxt['from_m']}"
                    f" (rows {prev['excel_row']}, {nxt['excel_row']})"
                )


def build_serial_chains(gate_ids: list[str]) -> list[list[str]]:
    """Maximal serial gate chains under the naming grammar — each chain is one
    canal chainage axis. Chain key = (parent-path tuples, branch index); members
    ordered by serial position. Each chain must start at position 0, but sparse
    later positions are retained because SCADA ids are physical identifiers, not
    list indexes to renumber when a nonexistent valve is removed."""
    chains: dict = {}
    for gate_id in gate_ids:
        tuples = parse_gate_id(gate_id)
        key = (tuple(tuples[:-1]), tuples[-1][0])
        chains.setdefault(key, []).append((tuples[-1][1], gate_id))
    result = []
    for key, members in chains.items():
        members.sort()
        positions = [pos for pos, _ in members]
        if len(positions) != len(set(positions)):
            raise WorkbookError(
                f"chain {key}: duplicate serial position in {positions}"
            )
        if positions[0] != 0:
            raise WorkbookError(
                f"chain {key}: serial positions {positions} must start at position 0"
            )
        result.append([gate_id for _, gate_id in members])
    return result


def validate_gate_distances(chains: list[list[str]], gates_by_id: dict) -> None:
    """Cross-check Sheet1 distances against authoritative gate chainages."""
    for chain in chains:
        for upstream, downstream in zip(chain, chain[1:]):
            gate = gates_by_id[upstream]
            expected_km = (gates_by_id[downstream]["km_m"] - gate["km_m"]) / 1000.0
            distance_km = gate["distance_km"]
            if (
                distance_km is None
                or abs(distance_km - expected_km) * 1000.0 > LENGTH_TOLERANCE_M
            ):
                raise WorkbookError(
                    f"gate {upstream}: Sheet1 distance {distance_km!r} km does not"
                    f" match chainage delta {expected_km:g} km to {downstream}"
                )


def match_chains_to_survey(
    chains: list[list[str]], canal_by_gate: dict, survey_canals: list[str]
) -> dict:
    """Survey canal -> gate chain, matched by member Sheet1 canal names, 1:1.

    Names compare whitespace-normalized. Zero or multiple candidate chains for
    a surveyed canal, or one chain claimed by two surveyed canals, all fail
    closed — a misassigned axis would hang real geometry on the wrong reaches.
    """

    def norm(name):
        return " ".join(str(name).split()) if name is not None else None

    matched: dict = {}
    claimed: dict = {}
    for canal in survey_canals:
        candidates = [
            chain
            for chain in chains
            if any(norm(canal_by_gate.get(g)) == norm(canal) for g in chain)
        ]
        if len(candidates) != 1:
            raise WorkbookError(
                f"survey canal {canal!r} matches {len(candidates)} gate chains"
                " — expected exactly 1"
            )
        chain = candidates[0]
        head = chain[0]
        if head in claimed:
            raise WorkbookError(
                f"gate chain starting {head!r} claimed by surveys"
                f" {claimed[head]!r} and {canal!r}"
            )
        claimed[head] = canal
        matched[canal] = chain
    return matched


def _subtract_intervals(span: tuple, holes: list) -> list:
    """Sub-intervals of ``span`` not covered by ``holes`` (sorted, disjoint)."""
    gaps = []
    cursor, end = span
    for lo, hi in sorted(holes):
        lo, hi = max(lo, cursor), min(hi, end)
        if hi <= lo:
            continue
        if lo > cursor + CHAINAGE_EPS:
            gaps.append((cursor, lo))
        cursor = max(cursor, hi)
    if end > cursor + CHAINAGE_EPS:
        gaps.append((cursor, end))
    return gaps


def assign_rows_to_edges(positions: list, rows: list) -> tuple:
    """Cut survey rows into gate-to-gate pieces along one canal axis.

    ``positions``: ordered (gate_id, km_m), strictly increasing — the axis.
    Returns (pieces_by_edge, skipped_rows, edge_coverage):
    - pieces_by_edge: {(upstream, downstream): [{'from_m','to_m','row'}]} —
      rows clipped at gate positions, chainage arithmetic only;
    - skipped_rows: spans that cannot be emitted, each with a reason
      (missing_cross_section / beyond_last_gate / before_first_gate);
    - edge_coverage: per edge covered_m/gap_m/status and each missing span
      with its reason — measured, never defaulted.
    """
    for (_, a), (_, b) in zip(positions, positions[1:]):
        if b <= a + CHAINAGE_EPS:
            raise WorkbookError(
                f"gate positions must be strictly increasing along the axis,"
                f" got {a} then {b}"
            )
    axis_start, axis_end = positions[0][1], positions[-1][1]

    def skip_reason(row):
        # A section without a full cross-section cannot load; one without its
        # design area or discharge would make consumers default the slope
        # divergently (0.0 vs 1.0) or silently lose the capacity bound — a
        # stale-formula workbook (saved without recalculation) presents exactly
        # this way. Report the row, never emit a partial one.
        if any(
            row.get(key) is None or row[key] <= 0
            for key in ("bottom_width_m", "depth_m", "manning_n", "bed_slope")
        ):
            return "missing_cross_section"
        if any(row.get(key) is None or row[key] <= 0 for key in ("area_m2", "qd")):
            return "missing_design_values"
        return None

    skipped = []
    usable = []
    for row in rows:
        reason = skip_reason(row)
        if reason is not None:
            skipped.append(
                {
                    "reason": reason,
                    "canal": row["canal"],
                    "from_m": row["from_m"],
                    "to_m": row["to_m"],
                    "excel_row": row["excel_row"],
                }
            )
            continue
        for lo, hi, reason in (
            (row["from_m"], min(row["to_m"], axis_start), "before_first_gate"),
            (max(row["from_m"], axis_end), row["to_m"], "beyond_last_gate"),
        ):
            if hi > lo + CHAINAGE_EPS:
                skipped.append(
                    {
                        "reason": reason,
                        "canal": row["canal"],
                        "from_m": lo,
                        "to_m": hi,
                        "excel_row": row["excel_row"],
                    }
                )
        if min(row["to_m"], axis_end) > max(row["from_m"], axis_start) + CHAINAGE_EPS:
            usable.append(row)

    unemittable_spans = [
        (s["from_m"], s["to_m"], s["reason"])
        for s in skipped
        if s["reason"] in ("missing_cross_section", "missing_design_values")
    ]
    pieces_by_edge: dict = {}
    coverage = []
    for (up_id, up_m), (down_id, down_m) in zip(positions, positions[1:]):
        pieces = []
        for row in usable:
            lo, hi = max(row["from_m"], up_m), min(row["to_m"], down_m)
            if hi > lo + CHAINAGE_EPS:
                pieces.append({"from_m": lo, "to_m": hi, "row": row})
        pieces.sort(key=lambda p: p["from_m"])
        if pieces:
            pieces_by_edge[(up_id, down_id)] = pieces
        covered = sum(p["to_m"] - p["from_m"] for p in pieces)
        span = down_m - up_m
        missing = []
        holes = [(lo, hi) for lo, hi, _ in unemittable_spans]
        for lo, hi in _subtract_intervals(
            (up_m, down_m), [(p["from_m"], p["to_m"]) for p in pieces]
        ):
            for sub_lo, sub_hi in _subtract_intervals((lo, hi), holes):
                missing.append((sub_lo, sub_hi, "unsurveyed"))
            for hole_lo, hole_hi, hole_reason in unemittable_spans:
                sub_lo, sub_hi = max(lo, hole_lo), min(hi, hole_hi)
                if sub_hi > sub_lo + CHAINAGE_EPS:
                    missing.append((sub_lo, sub_hi, f"survey_row_{hole_reason}"))
        missing.sort()
        status = (
            "full"
            if span - covered <= CHAINAGE_EPS
            else ("none" if covered <= CHAINAGE_EPS else "partial")
        )
        coverage.append(
            {
                "upstream": up_id,
                "downstream": down_id,
                "span_m": _intish(span),
                "covered_m": _intish(covered),
                "gap_m": _intish(span - covered),
                "segments": len(pieces),
                "status": status,
                "missing": [
                    {
                        "from_km": format_km_marker(lo),
                        "to_km": format_km_marker(hi),
                        "reason": reason,
                    }
                    for lo, hi, reason in missing
                ],
            }
        )
    return pieces_by_edge, skipped, coverage


def classify_edges(
    gate_ids: list,
    canal_by_gate: dict,
    km_by_gate: dict,
    matched_head_canals: dict,
    serial_coverage: dict,
) -> list:
    """One coverage entry per network edge, with fail-closed categories.

    - source: the root edge S -> M(0,0).
    - serial (measured): a gate-to-gate reach of a survey-matched chain — its
      coverage entry from assign_rows_to_edges.
    - junction_head: the child is a chain HEAD sitting at its offtake — zero
      chainage span, nothing to survey.
    - serial / status none: a reach of a MULTI-gate chain with no survey canal.
      A canal whose survey block disappears must show as a hole, never be
      absorbed into 'not_applicable' (that would certify the loss as normal).
    - offtake: a single-gate unmatched chain (Waste Way, FTOs) — a turnout with
      no canal of its own; its water runs in the parent canal's reaches.
    """
    gate_chain_head = {}
    single_gate_chains = set()
    for chain in build_serial_chains(gate_ids):
        for gate_id in chain:
            gate_chain_head[gate_id] = chain[0]
        if len(chain) == 1:
            single_gate_chains.add(chain[0])
    entries = []
    for parent, child in edges_from_names(gate_ids):
        head = gate_chain_head.get(child)
        if parent == ROOT:
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "source",
                "status": "not_applicable",
            }
        elif (parent, child) in serial_coverage:
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "serial",
                **{
                    k: v
                    for k, v in serial_coverage[(parent, child)].items()
                    if k not in ("upstream", "downstream")
                },
            }
        elif head == child and head in matched_head_canals:
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "junction_head",
                "status": "not_applicable",
                "canal": matched_head_canals[head],
            }
        elif head == child and head in single_gate_chains:
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "offtake",
                "status": "not_applicable",
                "canal": canal_by_gate.get(child),
                "km": km_by_gate.get(child),
            }
        elif head == child:
            # head of an UNSURVEYED multi-gate chain: the junction itself has no
            # span, but the canal's serial reaches below will show as holes.
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "junction_head",
                "status": "not_applicable",
                "canal": canal_by_gate.get(head),
            }
        else:
            entry = {
                "upstream": parent,
                "downstream": child,
                "category": "serial",
                "status": "none",
                "canal": canal_by_gate.get(head),
                "missing": [{"reason": "canal_not_in_survey"}],
            }
        entries.append(entry)
    return entries


def _metadata(workbook: Path, sha256: str, description: str) -> dict:
    return {
        "source_workbook": workbook.name,
        "source_sha256": sha256,
        "generator": GENERATOR,
        "description": description,
    }


def artifact_bytes(artifact: dict) -> bytes:
    """The exact bytes main() writes for an artifact — shared so lineage
    hashes always match the committed files."""
    return (json.dumps(artifact, indent=2, ensure_ascii=False) + "\n").encode(
        "utf-8"
    )


def calibration_similarity(target: dict, donor: dict) -> float:
    score = 0.0
    target_shape, donor_shape = target.get("shape"), donor.get("shape")
    if target_shape is not None and target_shape == donor_shape:
        score += SHAPE_SIMILARITY_WEIGHT
    dimension_ratios = []
    for key in ("width_m", "height_m"):
        target_value, donor_value = target.get(key), donor.get(key)
        if target_value is not None and donor_value is not None:
            dimension_ratios.append(
                min(target_value, donor_value) / max(target_value, donor_value)
            )
    if dimension_ratios:
        score += (
            DIMENSION_SIMILARITY_WEIGHT * sum(dimension_ratios) / len(dimension_ratios)
        )
    if target.get("canal_class") == donor.get("canal_class"):
        score += CANAL_CLASS_SIMILARITY_WEIGHT
    return round(score, 6)


def infer_calibration(target: dict, measured: list[dict], source_version: str) -> dict:
    target_shape = target.get("shape")
    if target_shape is not None:
        same_shape = [gate for gate in measured if gate.get("shape") == target_shape]
        candidates = same_shape or measured
    else:
        same_class = [
            gate
            for gate in measured
            if gate.get("canal_class") == target.get("canal_class")
        ]
        candidates = same_class or measured
    # Drop r2<=0 donors AFTER shape/class selection: such a donor carries zero
    # confidence and zero coefficient weight, yet the runtime loader rejects any
    # inference whose confidence is not below EVERY donor's (a zero-r2 donor makes
    # that impossible), so it must never reach ranking/lineage (2.3-retro MEDIUM).
    # Filtering here — not before selection — preserves fail-closed behavior: a
    # target whose only same-shape donor has r2<=0 raises below instead of silently
    # borrowing a differently-shaped donor's flow law.
    candidates = [gate for gate in candidates if (gate.get("r2") or 0) > 0]
    ranked = sorted(
        ((calibration_similarity(target, donor), donor) for donor in candidates),
        key=lambda item: (-item[0], item[1]["gate_id"]),
    )
    selected = [
        (score, donor) for score, donor in ranked[:MAX_SIMILAR_GATES] if score > 0
    ]
    if not selected:
        raise WorkbookError(
            f"gate {target['gate_id']}: no similar measured donor with positive r2 available"
        )
    coefficient_weights = [score * donor["r2"] for score, donor in selected]
    total_weight = sum(coefficient_weights)
    if total_weight <= 0:
        # Every selected donor has score>0 and r2>0, but the product of a normal
        # score and a subnormal r2 can still underflow to 0.0; guard the division
        # fail-closed rather than dividing by zero.
        raise WorkbookError(
            f"gate {target['gate_id']}: donor coefficient weights underflowed to zero"
        )
    k1 = (
        sum(
            donor["k1"] * weight
            for weight, (_, donor) in zip(coefficient_weights, selected)
        )
        / total_weight
    )
    k2 = (
        sum(
            donor["k2"] * weight
            for weight, (_, donor) in zip(coefficient_weights, selected)
        )
        / total_weight
    )
    similarity_weighted_fit = total_weight / sum(score for score, _ in selected)
    confidence = similarity_weighted_fit * selected[0][0] * INFERENCE_CONFIDENCE_PENALTY
    return {
        "calibration_method": "inferred",
        "k1": round(k1, 6),
        "k2": round(k2, 6),
        "confidence": round(confidence, 6),
        "source_gate_ids": [donor["gate_id"] for _, donor in selected],
        "source_version": source_version,
    }


def build_gate_calibrations(
    gates: list[dict], workbook_path: Path, sha256: str
) -> dict:
    records = {}
    counts = {"measured": 0, "inferred": 0, "default": 0}
    measured = []
    for gate in gates:
        gate_id = gate["gate_id"]
        triplet = (gate.get("k1"), gate.get("k2"), gate.get("r2"))
        present = tuple(value is not None for value in triplet)
        if any(present) and not all(present):
            raise WorkbookError(
                f"gate {gate_id}: k1/k2/r2 must be all present or all blank"
            )
        if all(present):
            k1, k2, r2 = triplet
            if k1 <= 0 or k2 >= 0 or not 0.0 <= r2 <= 1.0:
                raise WorkbookError(
                    f"gate {gate_id}: invalid measured k1/k2/r2 {triplet!r}"
                )
            measured.append(gate)
    inference_version = f"similar-gate-v1:{sha256}"
    measured_gate_ids = {gate["gate_id"] for gate in measured}
    for gate in gates:
        gate_id = gate["gate_id"]
        if gate_id in measured_gate_ids:
            k1, k2, r2 = gate["k1"], gate["k2"], gate["r2"]
            record = {
                "gate_id": gate_id,
                "calibration_method": "measured",
                "k1": _intish(k1),
                "k2": _intish(k2),
                "confidence": _intish(r2),
                "source_gate_ids": [gate_id],
                "source_version": sha256,
            }
            counts["measured"] += 1
        else:
            record = {
                "gate_id": gate_id,
                **infer_calibration(gate, measured, inference_version),
            }
            counts["inferred"] += 1
        for source_key, output_key in (
            ("canal_class", "canal_class"),
            ("shape", "shape"),
            ("width_m", "width_m"),
            ("height_m", "height_m"),
            ("q_max", "q_max_m3s"),
            ("zone", "zone"),
        ):
            if gate.get(source_key) is not None:
                record[output_key] = (
                    _intish(gate[source_key])
                    if source_key not in ("shape", "canal_class")
                    else gate[source_key]
                )
        for structure_key in (
            "design_fsl_msl_m",
            "sill_msl_m",
            "structure_max_flow_m3s",
            "design_fsl_reference_side",
            "structure_data_status",
            "structure_role",
        ):
            value = gate[structure_key]
            record[structure_key] = (
                _intish(value) if isinstance(value, (int, float)) else value
            )
        records[gate_id] = record
    return {
        "metadata": {
            **_metadata(
                workbook_path,
                sha256,
                (
                    "Gate calibration coefficients and provenance from Sheet1;"
                    " unmeasured gates use provisional similar-gate inference."
                ),
            ),
            "total_gates": len(records),
            "gates_by_calibration_method": counts,
            "intended_use": "planning_only",
            "design_fsl_reference_side": DESIGN_FSL_REFERENCE_SIDE,
        },
        "gates": records,
    }


def build_all(workbook_path) -> dict:
    """The four canonical artifacts from one workbook, deterministically."""
    import openpyxl

    workbook_path = Path(workbook_path)
    sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    gates = extract_gates(wb["Sheet1"])
    survey = extract_survey_rows(wb["Characteristics"])
    validate_survey_rows(survey)

    gate_ids = [g["gate_id"] for g in gates]
    gates_by_id = {g["gate_id"]: g for g in gates}
    canal_by_gate = {g["gate_id"]: g["canal"] for g in gates}
    chains = build_serial_chains(gate_ids)
    validate_gate_distances(chains, gates_by_id)
    survey_canals = list(dict.fromkeys(row["canal"] for row in survey))
    matched = match_chains_to_survey(chains, canal_by_gate, survey_canals)
    chain_canal = {chain[0]: canal for canal, chain in matched.items()}

    # --- per-canal mapping ------------------------------------------------
    rows_by_canal: dict = {}
    for row in survey:
        rows_by_canal.setdefault(row["canal"], []).append(row)
    sections = []
    reaches = []
    serial_coverage: dict = {}
    all_skipped = []
    chain_order = {chain[0]: i for i, chain in enumerate(chains)}
    for canal in sorted(matched, key=lambda c: chain_order[matched[c][0]]):
        chain = matched[canal]
        positions = [(gid, gates_by_id[gid]["km_m"]) for gid in chain]
        pieces_by_edge, skipped, coverage = assign_rows_to_edges(
            positions, rows_by_canal[canal]
        )
        all_skipped.extend(skipped)
        for entry, (up_id, up_m), (_, down_m) in zip(
            coverage, positions, positions[1:]
        ):
            entry["canal"] = canal
            serial_coverage[(entry["upstream"], entry["downstream"])] = entry
            # The gate-to-gate span every runtime consumer needs: head/tail gap
            # measurement (reach_chainage_gap_m) and the legacy solver's
            # physical reach length.
            reaches.append(
                {
                    "from_node": entry["upstream"],
                    "to_node": entry["downstream"],
                    "canal_name": canal,
                    "from_km": format_km_marker(up_m),
                    "to_km": format_km_marker(down_m),
                    "span_m": entry["span_m"],
                    "covered_m": entry["covered_m"],
                    "gap_m": entry["gap_m"],
                    "segments": entry["segments"],
                }
            )
        section_no = 0
        for edge_key in zip(chain, chain[1:]):
            for piece in pieces_by_edge.get(edge_key, []):
                section_no += 1
                row = piece["row"]
                up, down = edge_key
                zone = gates_by_id[down]["zone"]
                if zone is None:
                    zone = gates_by_id[up]["zone"]
                if zone is None:
                    raise WorkbookError(f"edge {edge_key}: neither gate carries a zone")
                # area_m2 and qd are emission-required (skip_reason), so
                # side_slope and q_max are present on EVERY section — consumers
                # never fall back to a divergent default slope or lose a bound.
                hydraulic_params = {
                    "manning_n": row["manning_n"],
                    "bed_slope": row["bed_slope"],
                }
                lining = lining_from_note(row["note"])
                if lining is not None:
                    hydraulic_params["lining_type"] = lining
                hydraulic_params["q_max"] = _intish(row["qd"])
                if row.get("q_rotation") is not None and row["q_rotation"] > 0:
                    # rotation-plan Qmax, retained so no capacity information is
                    # deleted (binding stays Qd pending the 2.3 calibration PR).
                    hydraulic_params["q_rotation_plan"] = _intish(row["q_rotation"])
                sections.append(
                    {
                        "from_node": up,
                        "to_node": down,
                        "canal_name": canal,
                        "zone": zone,
                        "section_no": section_no,
                        "from_km": format_km_marker(piece["from_m"]),
                        "to_km": format_km_marker(piece["to_m"]),
                        "geometry": {
                            "length_m": _intish(piece["to_m"] - piece["from_m"]),
                            "cross_section": {
                                "type": "trapezoidal",
                                "bottom_width_m": _intish(row["bottom_width_m"]),
                                "depth_m": _intish(row["depth_m"]),
                                "side_slope": derive_side_slope(
                                    row["area_m2"],
                                    row["bottom_width_m"],
                                    row["depth_m"],
                                ),
                            },
                            "hydraulic_params": hydraulic_params,
                        },
                    }
                )

    # --- network.json -----------------------------------------------------
    network_gates = {}
    gate_chain_head = {}
    for chain in chains:
        for gate_id in chain:
            gate_chain_head[gate_id] = chain[0]
    for order, gate in enumerate(gates):
        canal = gate["canal"]
        if canal is None:
            canal = chain_canal.get(gate_chain_head[gate["gate_id"]])
        entry = {
            "canal": canal,
            "zone": gate["zone"],
            "parsed": gate["parsed"],
            "order": order,
            "q_max": gate["q_max"],
            "area": gate["area"],
            "km": gate["km_text"],
        }
        if gate["zone"] is None:
            # zone: null breaks legacy `gate.get('zone', 0)` consumers (the key
            # existing returns None); absence keeps their default and is honest.
            del entry["zone"]
        network_gates[gate["gate_id"]] = entry
    edges = [list(edge) for edge in edges_from_names(gate_ids)]
    network = {
        "metadata": {
            **_metadata(
                workbook_path,
                sha256,
                (
                    "Canonical network: gates from Sheet1; edges regenerated"
                    " deterministically from the gate-id naming grammar"
                    " (core.network_topology.edges_from_names, F-11b)."
                ),
            ),
            "total_gates": len(network_gates),
            "total_connections": len(edges),
            "canonical": True,
        },
        "gates": network_gates,
        "edges": edges,
    }

    # --- canal_geometry.json ----------------------------------------------
    by_zone: dict = {}
    for section in sections:
        key = f"zone_{section['zone']}"
        by_zone[key] = by_zone.get(key, 0) + 1
    canal_geometry = {
        "metadata": {
            **_metadata(
                workbook_path,
                sha256,
                (
                    "Surveyed canal geometry (Characteristics sheet) cut into"
                    " gate-to-gate reach segments at Sheet1 gate chainages"
                    " (multi-segment reach model, Wave 2.1a/2.1b)."
                ),
            ),
            "notes": {
                "q_max": "design discharge Qd (ม.3/วินาที)",
                "side_slope": "derived from design flow area: z = (A/D - B)/D",
                "lining_type": "from the survey note; absent = unknown lining",
                "coverage": "see geometry_coverage.json for unmapped chainage",
            },
        },
        "canal_sections": sections,
        "reaches": reaches,
        "summary": {
            "total_sections": len(sections),
            "by_zone": by_zone,
        },
    }

    # --- geometry_coverage.json --------------------------------------------
    coverage_edges = classify_edges(
        gate_ids=gate_ids,
        canal_by_gate=canal_by_gate,
        km_by_gate={g["gate_id"]: g["km_text"] for g in gates},
        matched_head_canals=chain_canal,
        serial_coverage=serial_coverage,
    )
    by_status: dict = {}
    by_category: dict = {}
    for entry in coverage_edges:
        by_status[entry["status"]] = by_status.get(entry["status"], 0) + 1
        by_category[entry["category"]] = by_category.get(entry["category"], 0) + 1
    skipped_rows: dict = {}
    for skip in all_skipped:
        skipped_rows[skip["reason"]] = skipped_rows.get(skip["reason"], 0) + 1
    geometry_coverage = {
        "metadata": _metadata(
            workbook_path,
            sha256,
            (
                "Per-edge survey coverage of the canonical network: which reaches"
                " carry geometry, which chainage is unmapped and why. Junction"
                " heads sit at their offtake (zero chainage span); offtakes"
                " (Waste Way, FTOs) have no surveyed canal of their own."
            ),
        ),
        "edges": coverage_edges,
        "summary": {
            "total_edges": len(coverage_edges),
            "by_status": by_status,
            "by_category": by_category,
            "sections_emitted": len(sections),
            "survey_rows": len(survey),
            "skipped_rows": skipped_rows,
            "skipped_row_detail": [
                {
                    "reason": s["reason"],
                    "canal": s["canal"],
                    "from_km": format_km_marker(s["from_m"]),
                    "to_km": format_km_marker(s["to_m"]),
                    "excel_row": s["excel_row"],
                }
                for s in all_skipped
            ],
        },
    }
    gate_calibrations = build_gate_calibrations(gates, workbook_path, sha256)
    routing = derive_routing_topology(network, geometry_coverage, canal_geometry)
    routing_payload = routing_topology_payload(routing)
    role_counts: dict[str, int] = {}
    for element in routing.elements:
        role_counts[element.role.value] = role_counts.get(element.role.value, 0) + 1
    routing_topology = {
        "metadata": _metadata(
            workbook_path,
            sha256,
            "Derived typed hydraulic-routing topology: boundary/transport/"
            "branch-structure/withdrawal-structure elements over the canonical "
            "network, with the single virtual junction J(LMC,0+170) splitting "
            "the composite M(0,0)->M(0,2) span. Generated — never hand-edit.",
        ),
        "schema_version": routing_payload["schema_version"],
        "root_node_id": ROOT,
        "lineage": {
            "source_artifacts": [
                {
                    "source_id": name,
                    "filename": f"{name}.json",
                    "sha256": hashlib.sha256(
                        artifact_bytes(artifact)
                    ).hexdigest(),
                }
                for name, artifact in (
                    ("network", network),
                    ("geometry_coverage", geometry_coverage),
                    ("canal_geometry", canal_geometry),
                )
            ],
        },
        "elements": routing_payload["elements"],
        "summary": {
            "total_nodes": len(
                {ROOT}
                | {element.downstream_node_id for element in routing.elements}
            ),
            "total_elements": len(routing.elements),
            "by_role": role_counts,
        },
        "content_hash": routing.content_hash,
    }
    return {
        "network": network,
        "canal_geometry": canal_geometry,
        "geometry_coverage": geometry_coverage,
        "gate_calibrations": gate_calibrations,
        "routing_topology": routing_topology,
    }


def main(argv=None) -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    parser.add_argument("workbook", nargs="?", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    args = parser.parse_args(argv)
    artifacts = build_all(args.workbook)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    for name in (
        "network",
        "canal_geometry",
        "geometry_coverage",
        "gate_calibrations",
        "routing_topology",
    ):
        path = out_dir / f"{name}.json"
        path.write_bytes(artifact_bytes(artifacts[name]))
        print(f"wrote {path}")
    summary = artifacts["geometry_coverage"]["summary"]
    print(
        f"sections: {summary['sections_emitted']}"
        f" | edge status: {summary['by_status']}"
        f" | skipped rows: {summary['skipped_rows']}"
    )


if __name__ == "__main__":
    main()
